from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pydantic import Field

from ard_ossie.docling_parser import Evidence
from ard_ossie.ingestion import SourceRole
from ard_ossie.models import Sha256, StrictModel


class DictionaryColumn(StrictModel):
    ordinal: int = Field(gt=0)
    name: str
    logical_name: str | None = None
    data_type: str
    nullable: bool
    primary_key: bool
    foreign_key: str | None = None
    description: str | None = None
    formula: str | None = None
    comment: str | None = None
    evidence: Evidence


class DictionaryTable(StrictModel):
    locator: str
    description: str | None = Field(default=None, exclude_if=lambda value: value is None)
    columns: list[DictionaryColumn]


class ParsedDictionary(StrictModel):
    source_hash: Sha256
    tables: list[DictionaryTable]


_REQUIRED_HEADERS = frozenset(
    {"platform", "catalog", "schema", "table", "column", "data_type", "nullable", "pk"}
)
_KOREAN_REQUIRED_HEADERS = frozenset({"column", "description", "data_type", "key", "nullable"})
_KOREAN_METADATA_LABELS = {
    "저장_플랫폼_및_세부_위치": "location",
    "테이블_명": "table",
    "테이블_설명": "description",
}


def parse_dictionary(path: str | Path, *, source_hash: str) -> ParsedDictionary:
    workbook = load_workbook(Path(path), data_only=False, read_only=False)
    grouped: dict[str, list[DictionaryColumn]] = defaultdict(list)
    descriptions: dict[str, str | None] = {}
    parsed_sheet = False
    try:
        for sheet in workbook.worksheets:
            tables = _parse_flat_sheet(sheet, source_hash=source_hash)
            if tables is None:
                tables = _parse_korean_template_sheet(sheet, source_hash=source_hash)
            if tables is None:
                continue
            parsed_sheet = True
            for table in tables:
                previous_description = descriptions.get(table.locator)
                if (
                    previous_description is not None
                    and table.description is not None
                    and previous_description != table.description
                ):
                    raise ValueError(f"CONFLICTING_DICTIONARY_TABLE_DESCRIPTION: {table.locator}")
                descriptions[table.locator] = previous_description or table.description
                for column in table.columns:
                    grouped[table.locator].append(
                        column.model_copy(update={"ordinal": len(grouped[table.locator]) + 1})
                    )
    finally:
        workbook.close()

    if not parsed_sheet:
        raise ValueError("MISSING_DICTIONARY_HEADERS")
    for locator, columns in grouped.items():
        _validate_unique_columns(locator, columns)
    return ParsedDictionary(
        source_hash=source_hash,
        tables=[
            DictionaryTable(
                locator=locator,
                description=descriptions.get(locator),
                columns=grouped[locator],
            )
            for locator in sorted(grouped)
        ],
    )


def _parse_flat_sheet(sheet, *, source_hash: str) -> list[DictionaryTable] | None:
    headers = {
        _normalize_header(cell.value): index
        for index, cell in enumerate(sheet[1], start=1)
        if cell.value is not None
    }
    if not headers or _REQUIRED_HEADERS - set(headers):
        return None
    grouped: dict[str, list[DictionaryColumn]] = defaultdict(list)
    for row_number in range(2, sheet.max_row + 1):
        column_name = _cell_value(sheet, row_number, headers["column"])
        if column_name is None or not str(column_name).strip():
            continue
        platform = _required_value(sheet, row_number, headers, "platform")
        catalog = _required_value(sheet, row_number, headers, "catalog")
        schema = _required_value(sheet, row_number, headers, "schema")
        table = _required_value(sheet, row_number, headers, "table")
        locator = "|".join(
            str(value).strip().lower() for value in (platform, catalog, schema, table)
        )
        fk_table = _as_optional_text(_optional_value(sheet, row_number, headers, "fk_table"))
        fk_column = _as_optional_text(_optional_value(sheet, row_number, headers, "fk_column"))
        if bool(fk_table) != bool(fk_column):
            raise ValueError(f"INCOMPLETE_FOREIGN_KEY: {sheet.title}!{row_number}")
        foreign_key = f"{fk_table}.{fk_column}" if fk_table and fk_column else None
        row_cells = sheet[row_number]
        start = row_cells[0].coordinate
        end = row_cells[len(headers) - 1].coordinate
        source_cell = sheet.cell(row=row_number, column=headers["column"])
        explicit_comment = _as_optional_text(
            _optional_value(sheet, row_number, headers, "comment")
        )
        grouped[locator].append(
            DictionaryColumn(
                ordinal=len(grouped[locator]) + 1,
                name=str(column_name).strip(),
                logical_name=_as_optional_text(
                    _optional_value(sheet, row_number, headers, "logical_name")
                ),
                data_type=str(_required_value(sheet, row_number, headers, "data_type")).strip(),
                nullable=_as_bool(_required_value(sheet, row_number, headers, "nullable")),
                primary_key=_as_bool(_required_value(sheet, row_number, headers, "pk")),
                foreign_key=foreign_key,
                description=_as_optional_text(
                    _optional_value(sheet, row_number, headers, "description")
                ),
                formula=_as_optional_text(
                    _optional_value(sheet, row_number, headers, "formula")
                ),
                comment=(
                    explicit_comment
                    if explicit_comment is not None
                    else (source_cell.comment.text if source_cell.comment else None)
                ),
                evidence=Evidence(
                    source_hash=source_hash,
                    role=SourceRole.DICTIONARY_EXCEL,
                    locator={"sheet": sheet.title, "range": f"{start}:{end}"},
                    excerpt=str(column_name).strip(),
                ),
            )
        )
    return [
        DictionaryTable(locator=locator, columns=columns)
        for locator, columns in grouped.items()
    ]


def _parse_korean_template_sheet(sheet, *, source_hash: str) -> list[DictionaryTable] | None:
    detected = _find_korean_header(sheet)
    if detected is None:
        return None
    header_row, headers, data_end_column = detected
    metadata = _find_korean_metadata(sheet, end_row=header_row - 1)
    if not metadata:
        return None
    missing_metadata = sorted(
        name
        for name in {"location", "table"}
        if name not in metadata or _is_blank(metadata[name])
    )
    if missing_metadata:
        raise ValueError(
            f"MISSING_KOREAN_DICTIONARY_METADATA: {sheet.title}:{missing_metadata[0]}"
        )
    locator = _korean_locator(metadata["location"], metadata["table"], sheet.title)
    columns: list[DictionaryColumn] = []
    data_start_column = min(headers.values())
    for row_number in range(header_row + 1, sheet.max_row + 1):
        values = [
            _cell_value(sheet, row_number, column)
            for column in range(data_start_column, data_end_column + 1)
        ]
        if all(_is_blank(value) for value in values):
            break
        column_name = _required_value(sheet, row_number, headers, "column")
        key_value = _as_optional_text(_optional_value(sheet, row_number, headers, "key"))
        key_tokens = set() if key_value is None else set(re.findall(r"[A-Z]+", key_value.upper()))
        start = f"{get_column_letter(data_start_column)}{row_number}"
        end = f"{get_column_letter(data_end_column)}{row_number}"
        columns.append(
            DictionaryColumn(
                ordinal=len(columns) + 1,
                name=str(column_name).strip(),
                data_type=str(_required_value(sheet, row_number, headers, "data_type")).strip(),
                nullable=_as_bool(_required_value(sheet, row_number, headers, "nullable")),
                primary_key="PK" in key_tokens,
                foreign_key=None,
                description=_as_optional_text(
                    _optional_value(sheet, row_number, headers, "description")
                ),
                evidence=Evidence(
                    source_hash=source_hash,
                    role=SourceRole.DICTIONARY_EXCEL,
                    locator={"sheet": sheet.title, "range": f"{start}:{end}"},
                    excerpt=str(column_name).strip(),
                ),
            )
        )
    if not columns:
        raise ValueError(f"EMPTY_KOREAN_DICTIONARY_TABLE: {sheet.title}")
    return [
        DictionaryTable(
            locator=locator,
            description=_as_optional_text(metadata.get("description")),
            columns=columns,
        )
    ]


def _find_korean_header(sheet) -> tuple[int, dict[str, int], int] | None:
    for row_number in range(1, min(sheet.max_row, 50) + 1):
        headers: dict[str, int] = {}
        data_end_column: int | None = None
        for cell in sheet[row_number]:
            if cell.value is None:
                continue
            normalized = _normalize_header(cell.value)
            if normalized == "컬럼명":
                headers["column"] = cell.column
            elif normalized.startswith("컬럼_설명"):
                headers["description"] = cell.column
            elif normalized == "type":
                headers["data_type"] = cell.column
            elif normalized == "key_여부":
                headers["key"] = cell.column
            elif normalized == "null_허용":
                headers["nullable"] = cell.column
            elif normalized == "파티션_기준_컬럼":
                data_end_column = cell.column
        if set(headers) >= _KOREAN_REQUIRED_HEADERS:
            return row_number, headers, data_end_column or max(headers.values())
    return None


def _find_korean_metadata(sheet, *, end_row: int) -> dict[str, object]:
    metadata: dict[str, object] = {}
    for row_number in range(1, end_row + 1):
        for cell in sheet[row_number]:
            name = _KOREAN_METADATA_LABELS.get(_normalize_header(cell.value))
            if name is None:
                continue
            value = _cell_value(sheet, row_number, cell.column + 1)
            metadata[name] = value
    return metadata


def _korean_locator(location: object, table: object, sheet_title: str) -> str:
    parts = [part.strip().lower() for part in str(location).split(".")]
    if len(parts) == 2:
        parts.insert(0, "unspecified")
    if len(parts) != 3 or any(not part or "|" in part for part in parts):
        raise ValueError(f"INVALID_KOREAN_DICTIONARY_LOCATION: {sheet_title}")
    table_name = str(table).strip().lower()
    if not table_name or "|" in table_name:
        raise ValueError(f"INVALID_KOREAN_DICTIONARY_TABLE: {sheet_title}")
    return "|".join((*parts, table_name))


def _validate_unique_columns(locator: str, columns: list[DictionaryColumn]) -> None:
    seen: set[str] = set()
    for column in columns:
        normalized = column.name.casefold()
        if normalized in seen:
            raise ValueError(f"DUPLICATE_DICTIONARY_COLUMN: {locator}:{column.name}")
        seen.add(normalized)


def _normalize_header(value: object) -> str:
    return re.sub(r"\s+", "_", str(value).replace("\xa0", " ").strip().lower())


def _is_blank(value: object) -> bool:
    return value is None or not str(value).strip()


def _cell_value(sheet, row: int, column: int):
    return sheet.cell(row=row, column=column).value


def _required_value(sheet, row: int, headers: dict[str, int], name: str):
    value = _cell_value(sheet, row, headers[name])
    if value is None or not str(value).strip():
        raise ValueError(f"MISSING_DICTIONARY_VALUE: {sheet.title}!{row}:{name}")
    return value


def _optional_value(sheet, row: int, headers: dict[str, int], name: str):
    column = headers.get(name)
    return None if column is None else _cell_value(sheet, row, column)


def _as_optional_text(value: object) -> str | None:
    if value is None or not str(value).strip():
        return None
    return str(value).strip()


def _as_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "y", "yes"}:
        return True
    if normalized in {"0", "false", "n", "no"}:
        return False
    raise ValueError(f"INVALID_BOOLEAN_VALUE: {value}")
