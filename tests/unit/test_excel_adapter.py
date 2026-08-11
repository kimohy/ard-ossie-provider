from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from ard_ossie.excel_adapter import parse_dictionary


def create_dictionary(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data Dictionary"
    sheet.append(
        [
            "Platform",
            "Catalog",
            "Schema",
            "Table",
            "Column",
            "Logical Name",
            "Data Type",
            "Nullable",
            "PK",
            "FK Table",
            "FK Column",
            "Description",
        ]
    )
    sheet.append(
        [
            "BigQuery",
            "analytics",
            "sales",
            "orders",
            "order_id",
            "Order ID",
            "INT64",
            "N",
            "Y",
            None,
            None,
            "Unique order identifier",
        ]
    )
    sheet.append(
        [
            "BigQuery",
            "analytics",
            "sales",
            "orders",
            "customer_id",
            "Customer ID",
            "INT64",
            "N",
            "N",
            "customers",
            "customer_id",
            "Ordering customer",
        ]
    )
    workbook.save(path)


def add_korean_template_sheet(
    workbook: Workbook,
    *,
    title: str,
    location: str,
    description: str,
    rows: list[tuple[str, str, str, str | None, str]],
    residue: tuple[int, tuple[str, str, str, str | None, str]] | None = None,
) -> None:
    if workbook.active.max_row == 1 and workbook.active["A1"].value is None:
        sheet = workbook.active
        sheet.title = title
    else:
        sheet = workbook.create_sheet(title)
    sheet["B3"] = "저장 플랫폼 및 세부 위치"
    sheet["C3"] = location
    sheet["B4"] = "테이블 명"
    sheet["C4"] = title
    sheet["B7"] = "테이블 설명"
    sheet["C7"] = description
    headers = [
        "컬럼명",
        "컬럼 설명\n(* Data Value의 단위/범위 설명이 필요한 경우 필수 추가 작성)",
        "Type",
        "Key 여부",
        "Null 허용",
        "PII 여부",
        "암호화 여부",
        "PII 태깅",
        "파티션 기준 컬럼",
    ]
    for column, value in enumerate(headers, start=2):
        sheet.cell(row=13, column=column, value=value)
    for row_number, values in enumerate(rows, start=14):
        for column, value in enumerate(values, start=2):
            sheet.cell(row=row_number, column=column, value=value)
    if residue is not None:
        row_number, values = residue
        for column, value in enumerate(values, start=2):
            sheet.cell(row=row_number, column=column, value=value)


def create_korean_template_dictionary(path: Path) -> None:
    workbook = Workbook()
    add_korean_template_sheet(
        workbook,
        title="marketing_campaign",
        location="synthetic_workspace.marketing_insight",
        description="가상 캠페인 합성 테이블",
        rows=[
            ("campaign_id", "가상 캠페인 식별자", "STRING", "PK", "N"),
            ("campaign_name", "가상 캠페인 명칭", "STRING", None, "N"),
            ("objective_type", "가상 목표 구분", "STRING", None, "Y"),
            ("channel_group", "추상 채널 그룹", "STRING", None, "Y"),
            ("market_zone", "가상 분석 권역", "STRING", None, "Y"),
            ("budget_units", "시뮬레이션 예산 단위", "NUMERIC", None, "Y"),
            ("start_date", "가상 캠페인 시작 일자", "DATE", None, "Y"),
            ("end_date", "가상 캠페인 종료 일자", "DATE", None, "Y"),
            ("campaign_status", "가상 캠페인 상태", "STRING", None, "Y"),
            ("priority_band", "가상 우선순위", "STRING", None, "Y"),
            ("planned_reach", "목표 노출 규모", "INTEGER", None, "Y"),
            ("planned_response", "목표 반응 규모", "INTEGER", None, "Y"),
            ("experiment_flag", "가상 실험 여부", "BOOLEAN", None, "Y"),
            ("source_created_at", "합성 레코드 생성 시각", "TIMESTAMP", None, "Y"),
            ("source_updated_at", "합성 레코드 갱신 시각", "TIMESTAMP", None, "Y"),
            ("loaded_at", "적재 시각", "TIMESTAMP", None, "N"),
        ],
        residue=(68, ("loaded_at", "테이블 생성 일시", "TIMESTAMP", None, "Y")),
    )
    add_korean_template_sheet(
        workbook,
        title="marketing_creative",
        location="synthetic_workspace.marketing_insight",
        description="가상 소재 합성 테이블",
        rows=[
            ("creative_id", "가상 소재 식별자", "STRING", "PK", "N"),
            ("campaign_id", "상위 캠페인 식별자", "STRING", "FK", "N"),
            ("format_type", "가상 소재 형식", "STRING", None, "Y"),
            ("message_theme", "가상 메시지 테마", "STRING", None, "Y"),
            ("variant_code", "가상 변형 코드", "STRING", None, "Y"),
            ("review_status", "가상 검토 상태", "STRING", None, "Y"),
            ("loaded_at", "적재 시각", "TIMESTAMP", None, "N"),
        ],
    )
    add_korean_template_sheet(
        workbook,
        title="marketing_delivery",
        location="synthetic_workspace.marketing_insight",
        description="가상 집행 합성 테이블",
        rows=[
            ("event_date", "가상 성과 기준 일자", "DATE", "PK", "N"),
            ("campaign_id", "가상 캠페인 식별자", "STRING", "PK, FK", "N"),
            ("creative_id", "가상 소재 식별자", "STRING", "PK, FK", "N"),
            ("audience_cluster", "가상 군집 라벨", "STRING", None, "Y"),
            ("impression_count", "가상 노출 수", "INTEGER", None, "Y"),
            ("engagement_count", "가상 반응 수", "INTEGER", None, "Y"),
            ("spend_units", "가상 비용 단위", "NUMERIC", None, "Y"),
            ("delivery_status", "가상 집행 상태", "STRING", None, "Y"),
            ("loaded_at", "적재 시각", "TIMESTAMP", None, "N"),
        ],
    )
    add_korean_template_sheet(
        workbook,
        title="marketing_outcome",
        location="synthetic_workspace.marketing_insight",
        description="가상 성과 합성 테이블",
        rows=[
            ("event_date", "가상 성과 기준 일자", "DATE", "PK", "N"),
            ("campaign_id", "가상 캠페인 식별자", "STRING", "PK, FK", "N"),
            ("interest_signal_count", "가상 관심 신호 수", "INTEGER", None, "Y"),
            ("action_signal_count", "가상 행동 신호 수", "INTEGER", None, "Y"),
            ("modeled_value_units", "가상 가치 단위", "NUMERIC", None, "Y"),
            ("attribution_window", "가상 관찰 창", "STRING", None, "Y"),
            ("quality_flag", "가상 품질 상태", "STRING", None, "Y"),
            ("loaded_at", "적재 시각", "TIMESTAMP", None, "N"),
        ],
        residue=(23, ("loaded_at", "테이블 생성 일시", "TIMESTAMP", None, "Y")),
    )
    workbook.save(path)


def test_excel_adapter_preserves_physical_schema_and_cell_evidence(tmp_path: Path) -> None:
    path = tmp_path / "dictionary.xlsx"
    create_dictionary(path)

    parsed = parse_dictionary(path, source_hash="a" * 64)

    assert len(parsed.tables) == 1
    table = parsed.tables[0]
    assert table.locator == "bigquery|analytics|sales|orders"
    assert [column.name for column in table.columns] == ["order_id", "customer_id"]
    assert table.columns[0].primary_key is True
    assert table.columns[1].foreign_key == "customers.customer_id"
    assert table.columns[0].evidence.locator == {
        "sheet": "Data Dictionary",
        "range": "A2:L2",
    }


def test_flat_dictionary_serialization_omits_absent_table_description(tmp_path: Path) -> None:
    path = tmp_path / "dictionary.xlsx"
    create_dictionary(path)

    payload = parse_dictionary(path, source_hash="a" * 64).model_dump(mode="json")

    assert "description" not in payload["tables"][0]


def test_excel_adapter_rejects_missing_required_header(tmp_path: Path) -> None:
    path = tmp_path / "dictionary.xlsx"
    workbook = Workbook()
    workbook.active.append(["Table", "Column"])
    workbook.save(path)

    try:
        parse_dictionary(path, source_hash="a" * 64)
    except ValueError as error:
        assert "MISSING_DICTIONARY_HEADERS" in str(error)
    else:
        raise AssertionError("missing headers must fail")


def test_excel_adapter_reads_explicit_formula_and_comment_columns(tmp_path: Path) -> None:
    path = tmp_path / "dictionary.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Platform",
            "Catalog",
            "Schema",
            "Table",
            "Column",
            "Data Type",
            "Nullable",
            "PK",
            "Formula",
            "Comment",
        ]
    )
    sheet.append(
        [
            "erp",
            "analytics",
            "sales",
            "orders",
            "net_amount",
            "DECIMAL",
            "false",
            "no",
            "gross_amount - tax_amount",
            "Validated finance rule",
        ]
    )
    workbook.save(path)

    column = parse_dictionary(path, source_hash="a" * 64).tables[0].columns[0]

    assert column.formula == "gross_amount - tax_amount"
    assert column.comment == "Validated finance rule"


def test_excel_adapter_rejects_unknown_boolean_text(tmp_path: Path) -> None:
    path = tmp_path / "dictionary.xlsx"
    create_dictionary(path)
    workbook = load_workbook(path)
    workbook.active["H2"] = "sometimes"
    workbook.save(path)
    workbook.close()

    with pytest.raises(ValueError, match="INVALID_BOOLEAN_VALUE"):
        parse_dictionary(path, source_hash="a" * 64)


def test_excel_adapter_rejects_half_specified_foreign_key(tmp_path: Path) -> None:
    path = tmp_path / "dictionary.xlsx"
    create_dictionary(path)
    workbook = load_workbook(path)
    workbook.active["K3"] = None
    workbook.save(path)
    workbook.close()

    with pytest.raises(ValueError, match="INCOMPLETE_FOREIGN_KEY"):
        parse_dictionary(path, source_hash="a" * 64)


def test_excel_adapter_parses_multi_sheet_korean_template_without_residue(
    tmp_path: Path,
) -> None:
    path = tmp_path / "dictionary.xlsx"
    create_korean_template_dictionary(path)

    parsed = parse_dictionary(path, source_hash="a" * 64)

    assert [table.locator for table in parsed.tables] == [
        "unspecified|synthetic_workspace|marketing_insight|marketing_campaign",
        "unspecified|synthetic_workspace|marketing_insight|marketing_creative",
        "unspecified|synthetic_workspace|marketing_insight|marketing_delivery",
        "unspecified|synthetic_workspace|marketing_insight|marketing_outcome",
    ]
    assert [len(table.columns) for table in parsed.tables] == [16, 7, 9, 8]
    assert sum(len(table.columns) for table in parsed.tables) == 40
    campaign, creative, delivery, outcome = parsed.tables
    assert campaign.description == "가상 캠페인 합성 테이블"
    assert campaign.columns[0].name == "campaign_id"
    assert campaign.columns[-1].name == "loaded_at"
    assert [column.ordinal for column in campaign.columns] == list(range(1, 17))
    assert campaign.columns[0].primary_key is True
    assert campaign.columns[0].nullable is False
    assert campaign.columns[0].evidence.locator == {
        "sheet": "marketing_campaign",
        "range": "B14:J14",
    }
    assert creative.columns[1].primary_key is False
    assert creative.columns[1].foreign_key is None
    assert creative.columns[4].nullable is True
    assert delivery.columns[1].primary_key is True
    assert delivery.columns[1].foreign_key is None
    assert outcome.columns[-1].name == "loaded_at"


def test_excel_adapter_accepts_explicit_platform_in_korean_location(tmp_path: Path) -> None:
    path = tmp_path / "dictionary.xlsx"
    workbook = Workbook()
    add_korean_template_sheet(
        workbook,
        title="marketing_campaign",
        location="bigquery.synthetic_workspace.marketing_insight",
        description="Campaigns",
        rows=[("campaign_id", "Identifier", "STRING", "PK", "N")],
    )
    workbook.save(path)

    parsed = parse_dictionary(path, source_hash="a" * 64)

    assert parsed.tables[0].locator == (
        "bigquery|synthetic_workspace|marketing_insight|marketing_campaign"
    )


def test_excel_adapter_rejects_malformed_korean_location(tmp_path: Path) -> None:
    path = tmp_path / "dictionary.xlsx"
    workbook = Workbook()
    add_korean_template_sheet(
        workbook,
        title="marketing_campaign",
        location="marketing_insight",
        description="Campaigns",
        rows=[("campaign_id", "Identifier", "STRING", "PK", "N")],
    )
    workbook.save(path)

    with pytest.raises(ValueError, match="INVALID_KOREAN_DICTIONARY_LOCATION"):
        parse_dictionary(path, source_hash="a" * 64)


def test_excel_adapter_rejects_duplicate_live_korean_columns(tmp_path: Path) -> None:
    path = tmp_path / "dictionary.xlsx"
    workbook = Workbook()
    add_korean_template_sheet(
        workbook,
        title="marketing_campaign",
        location="synthetic_workspace.marketing_insight",
        description="Campaigns",
        rows=[
            ("campaign_id", "Identifier", "STRING", "PK", "N"),
            ("CAMPAIGN_ID", "Duplicate", "STRING", None, "Y"),
        ],
    )
    workbook.save(path)

    with pytest.raises(ValueError, match="DUPLICATE_DICTIONARY_COLUMN"):
        parse_dictionary(path, source_hash="a" * 64)


def test_excel_adapter_rejects_incomplete_korean_sheet_in_mixed_workbook(
    tmp_path: Path,
) -> None:
    path = tmp_path / "dictionary.xlsx"
    create_dictionary(path)
    workbook = load_workbook(path)
    add_korean_template_sheet(
        workbook,
        title="marketing_campaign",
        location="",
        description="Campaigns",
        rows=[("campaign_id", "Identifier", "STRING", "PK", "N")],
    )
    workbook.save(path)
    workbook.close()

    with pytest.raises(ValueError, match="MISSING_KOREAN_DICTIONARY_METADATA"):
        parse_dictionary(path, source_hash="a" * 64)


def test_excel_adapter_ignores_header_only_non_template_sheet(tmp_path: Path) -> None:
    path = tmp_path / "dictionary.xlsx"
    create_dictionary(path)
    workbook = load_workbook(path)
    sheet = workbook.create_sheet("Korean Header Notes")
    for column, value in enumerate(
        ["컬럼명", "컬럼 설명", "Type", "Key 여부", "Null 허용"],
        start=2,
    ):
        sheet.cell(row=13, column=column, value=value)
    workbook.save(path)
    workbook.close()

    parsed = parse_dictionary(path, source_hash="a" * 64)

    assert [table.locator for table in parsed.tables] == [
        "bigquery|analytics|sales|orders"
    ]


def test_excel_adapter_rejects_korean_table_without_contiguous_data_rows(
    tmp_path: Path,
) -> None:
    path = tmp_path / "dictionary.xlsx"
    workbook = Workbook()
    add_korean_template_sheet(
        workbook,
        title="marketing_campaign",
        location="synthetic_workspace.marketing_insight",
        description="Campaigns",
        rows=[],
    )
    workbook.save(path)

    with pytest.raises(ValueError, match="EMPTY_KOREAN_DICTIONARY_TABLE"):
        parse_dictionary(path, source_hash="a" * 64)
